# Copyright (c) 2024 Finite State, Inc.
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Main report renderer that orchestrates all output formats."""

import json
import logging
import math
from datetime import UTC
from pathlib import Path
from typing import Any

import pandas as pd

from fs_report.models import Recipe, ReportData
from fs_report.renderers.csv_renderer import CSVRenderer
from fs_report.renderers.html_renderer import HTMLRenderer
from fs_report.renderers.md_renderer import MarkdownRenderer
from fs_report.renderers.xlsx_renderer import XLSXRenderer


class ReportRenderer:
    """Main renderer that coordinates all output formats."""

    def __init__(
        self, output_dir: str, config: Any = None, overwrite: bool = False
    ) -> None:
        """Initialize the report renderer."""
        self.output_dir = Path(output_dir)
        self.logger = logging.getLogger(__name__)
        self.config = config
        self.overwrite = overwrite

        # Initialize individual renderers
        self.csv_renderer = CSVRenderer()
        self.xlsx_renderer = XLSXRenderer()
        self.html_renderer = HTMLRenderer()
        self.md_renderer = MarkdownRenderer()

    def check_output_guard(self, recipe: Recipe) -> None:
        """Raise FileExistsError if the recipe output dir already has files and overwrite is off."""
        recipe_output_dir = self.output_dir / self._sanitize_filename(recipe.name)
        if recipe_output_dir.exists() and any(recipe_output_dir.iterdir()):
            if not self.overwrite:
                raise FileExistsError(
                    f"Output directory '{recipe_output_dir}' already contains files. "
                    "Use --overwrite to replace existing reports."
                )

    def render(self, recipe: Recipe, report_data: ReportData) -> list[str]:
        """Render reports in all configured formats. Returns a list of generated file paths."""
        self.logger.debug(f"Rendering reports for: {recipe.name}")

        # Create recipe-specific output directory
        recipe_output_dir = self.output_dir / self._sanitize_filename(recipe.name)
        recipe_output_dir.mkdir(parents=True, exist_ok=True)

        # Determine which formats to generate
        formats = getattr(recipe.output, "formats", None)
        if not formats:
            formats = ["csv", "xlsx", "html"]
        formats = [f.lower() for f in formats]

        # Auto-skip expensive formats for large datasets
        row_count = len(report_data.data) if hasattr(report_data.data, "__len__") else 0

        # When the transform returned a dict (chart/dashboard data), the
        # DataFrame in report_data.data is a CSV/XLSX fallback extracted from
        # the dict's "main" key.  The HTML template renders the compact chart
        # data, so its size is NOT proportional to the DataFrame row count.
        _additional = report_data.metadata.get("additional_data", {})
        _html_is_chart_driven = isinstance(
            (
                _additional.get("transform_result")
                if isinstance(_additional, dict)
                else None
            ),
            dict,
        )

        if row_count > 65_530 and "xlsx" in formats:
            formats = [f for f in formats if f != "xlsx"]
            self.logger.info(
                f"Skipping XLSX for {recipe.name}: {row_count:,} rows "
                f"exceed Excel's URL limit. Use CSV instead."
            )
        if row_count > 50_000 and "html" in formats and not _html_is_chart_driven:
            formats = [f for f in formats if f != "html"]
            self.logger.info(
                f"Skipping HTML for {recipe.name}: {row_count:,} rows "
                f"would produce an oversized file. Use CSV instead."
            )
            # Add CSV as fallback so we still produce output
            if "csv" not in formats:
                formats.append("csv")

        generated_files = []

        # Generate table-based formats if requested
        if any(fmt in formats for fmt in ["csv", "xlsx"]):
            generated_files += self._render_table_formats(
                recipe, report_data, recipe_output_dir, formats
            )

        # Generate HTML if requested
        if "html" in formats:
            generated_files += self._render_chart_formats(
                recipe, report_data, recipe_output_dir
            )

        # Generate JSON remediation package if requested
        if "json" in formats:
            generated_files += self._render_json(recipe, report_data, recipe_output_dir)

        # Generate PDF if requested
        if "pdf" in formats:
            generated_files += self._render_pdf(recipe, report_data, recipe_output_dir)

        # Generate Markdown agent prompt if requested
        if "md" in formats:
            generated_files += self._render_markdown(
                recipe, report_data, recipe_output_dir
            )

        return generated_files

    def _render_table_formats(
        self,
        recipe: Recipe,
        report_data: ReportData,
        output_dir: Path,
        formats: list[str],
    ) -> list[str]:
        """Render table-based formats (CSV, XLSX). Returns list of generated file paths."""
        generated_files = []
        try:
            # For CVA, use portfolio data instead of main data
            if "Component Vulnerability Analysis" in recipe.name:
                self.logger.debug(
                    "Using project-level data for Component Vulnerability Analysis table"
                )
                table_data = report_data.metadata.get(
                    "portfolio_data", report_data.data
                )
            else:
                table_data = report_data.data

            additional_data = report_data.metadata.get("additional_data", {})
            detail_findings = additional_data.get("detail_findings")
            detail_findings_churn = additional_data.get("detail_findings_churn")
            detail_component_churn = additional_data.get("detail_component_churn")
            has_detail = recipe.name == "Version Comparison" and (
                detail_findings is not None
                or detail_findings_churn is not None
                or detail_component_churn is not None
            )
            # Build the VC progression DataFrame (always, for all VC runs).
            vc_progression_df: pd.DataFrame | None = None
            if recipe.name == "Version Comparison":
                vc_projects = additional_data.get("projects", [])
                if vc_projects:
                    rows: list[dict] = []
                    for proj in vc_projects:
                        proj_name = proj.get("project_name", "")
                        for step in proj.get("progression", []):
                            row = dict(step)
                            row["project"] = proj_name
                            rows.append(row)
                    if rows:
                        vc_progression_df = pd.DataFrame(rows)
                        # Derive user-facing status column from fetch_failed flag.
                        if "fetch_failed" in vc_progression_df.columns:
                            vc_progression_df["status"] = vc_progression_df[
                                "fetch_failed"
                            ].apply(
                                lambda v: (
                                    "fetch_failed"
                                    if (v is True or v == True)  # noqa: E712
                                    else "ok"
                                )
                            )
                        else:
                            vc_progression_df["status"] = "ok"
                        # Drop internal boolean flags (superseded by status).
                        for _drop_col in ("fetch_failed", "delta_unavailable"):
                            if _drop_col in vc_progression_df.columns:
                                vc_progression_df = vc_progression_df.drop(
                                    columns=[_drop_col]
                                )
                        # Reorder: project, version, status, then everything else.
                        _cols = list(vc_progression_df.columns)
                        for _anchor in ("status", "version", "project"):
                            if _anchor in _cols:
                                _cols.remove(_anchor)
                                _cols.insert(0, _anchor)
                        vc_progression_df = vc_progression_df[_cols]
            if has_detail:
                detail_findings_df = (
                    detail_findings
                    if isinstance(detail_findings, pd.DataFrame)
                    else pd.DataFrame()
                )
                detail_findings_churn_df = (
                    detail_findings_churn
                    if isinstance(detail_findings_churn, pd.DataFrame)
                    else pd.DataFrame()
                )
                detail_component_churn_df = (
                    detail_component_churn
                    if isinstance(detail_component_churn, pd.DataFrame)
                    else pd.DataFrame()
                )

            # Scan Quality: summary + detail tables
            scan_quality_detail = additional_data.get("detail_table")
            has_scan_quality_detail = (
                recipe.name == "Scan Quality" and scan_quality_detail is not None
            )
            if has_scan_quality_detail:
                scan_quality_detail_df = (
                    scan_quality_detail
                    if isinstance(scan_quality_detail, pd.DataFrame)
                    else pd.DataFrame()
                )

            # Generate main files
            base_filename = self._sanitize_filename(recipe.name)

            # CSV output
            if "csv" in formats:
                csv_path = output_dir / f"{base_filename}.csv"
                self.csv_renderer.render(table_data, csv_path)
                self.logger.debug(f"Generated CSV: {csv_path}")
                generated_files.append(str(csv_path))
                if has_detail and not detail_findings_df.empty:
                    detail_csv = output_dir / f"{base_filename}_Detail_Findings.csv"
                    self.csv_renderer.render(detail_findings_df, detail_csv)
                    generated_files.append(str(detail_csv))
                if has_detail and not detail_findings_churn_df.empty:
                    findings_churn_csv = (
                        output_dir / f"{base_filename}_Detail_Findings_Churn.csv"
                    )
                    self.csv_renderer.render(
                        detail_findings_churn_df, findings_churn_csv
                    )
                    generated_files.append(str(findings_churn_csv))
                if has_detail and not detail_component_churn_df.empty:
                    churn_csv = (
                        output_dir / f"{base_filename}_Detail_Component_Churn.csv"
                    )
                    self.csv_renderer.render(detail_component_churn_df, churn_csv)
                    generated_files.append(str(churn_csv))
                if has_scan_quality_detail and not scan_quality_detail_df.empty:
                    detail_csv = output_dir / f"{base_filename}_Detail.csv"
                    self.csv_renderer.render(scan_quality_detail_df, detail_csv)
                    generated_files.append(str(detail_csv))
                if vc_progression_df is not None and not vc_progression_df.empty:
                    progression_csv = output_dir / f"{base_filename}_Progression.csv"
                    self.csv_renderer.render(vc_progression_df, progression_csv)
                    self.logger.debug(
                        f"Generated VC Progression CSV: {progression_csv}"
                    )
                    generated_files.append(str(progression_csv))
            # VC Progression XLSX (separate file, uses openpyxl for PatternFill support)
            if (
                "xlsx" in formats
                and vc_progression_df is not None
                and not vc_progression_df.empty
            ):
                progression_xlsx = output_dir / f"{base_filename}_Progression.xlsx"
                self._write_vc_progression_xlsx(vc_progression_df, progression_xlsx)
                self.logger.debug(f"Generated VC Progression XLSX: {progression_xlsx}")
                generated_files.append(str(progression_xlsx))
            # XLSX output
            if "xlsx" in formats:
                xlsx_path = output_dir / f"{base_filename}.xlsx"
                if has_detail:
                    sheets = [("Summary", table_data)]
                    if detail_findings_df is not None and not detail_findings_df.empty:
                        sheets.append(("Findings Detail", detail_findings_df))
                    if (
                        detail_findings_churn_df is not None
                        and not detail_findings_churn_df.empty
                    ):
                        sheets.append(("Findings Churn", detail_findings_churn_df))
                    if (
                        detail_component_churn_df is not None
                        and not detail_component_churn_df.empty
                    ):
                        sheets.append(("Component Churn", detail_component_churn_df))
                    self.xlsx_renderer.render_multi_sheet(sheets, xlsx_path)
                elif recipe.name == "Executive Dashboard":
                    # Executive Dashboard: multi-sheet with Project Summary + Top Risk Products
                    transform_result = additional_data.get("transform_result", {})
                    if isinstance(transform_result, dict):
                        project_table_data = transform_result.get("project_table")
                        top_risk_data = transform_result.get("top_risk_products")
                        ed_sheets: list[tuple[str, Any]] = []
                        if project_table_data:
                            ed_sheets.append(
                                ("Project Summary", pd.DataFrame(project_table_data))
                            )
                        if top_risk_data:
                            ed_sheets.append(
                                ("Top Risk Products", pd.DataFrame(top_risk_data))
                            )
                        if ed_sheets:
                            self.xlsx_renderer.render_multi_sheet(ed_sheets, xlsx_path)
                        else:
                            self.xlsx_renderer.render(
                                table_data, xlsx_path, recipe.name
                            )
                    else:
                        self.xlsx_renderer.render(table_data, xlsx_path, recipe.name)
                elif recipe.name == "Component List":
                    # Component List: multi-sheet with Summary + Detail
                    transform_result = additional_data.get("transform_result", {})
                    comp_summary = (
                        transform_result.get("component_summary")
                        if isinstance(transform_result, dict)
                        else None
                    )
                    if comp_summary and comp_summary.get("total_components", 0) > 0:
                        summary_rows = []
                        summary_rows.append(
                            {
                                "Metric": "Total Components",
                                "Value": comp_summary["total_components"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Unique Licenses",
                                "Value": comp_summary["unique_licenses"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "No License",
                                "Value": comp_summary["no_license_count"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Policy Violations",
                                "Value": comp_summary["violation_count"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Policy Warnings",
                                "Value": comp_summary["warning_count"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Copyleft (Strong)",
                                "Value": comp_summary["copyleft_strong"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Copyleft (Weak)",
                                "Value": comp_summary["copyleft_weak"],
                            }
                        )
                        summary_rows.append(
                            {
                                "Metric": "Permissive",
                                "Value": comp_summary["copyleft_permissive"],
                            }
                        )
                        summary_df = pd.DataFrame(summary_rows)

                        cl_sheets: list[tuple[str, Any]] = [
                            ("Summary", summary_df),
                            ("Detail", table_data),
                        ]

                        # Add distribution tables as additional sheets
                        lic_dist = comp_summary.get("license_distribution")
                        if isinstance(lic_dist, pd.DataFrame) and not lic_dist.empty:
                            cl_sheets.append(("License Distribution", lic_dist))

                        pol_dist = comp_summary.get("policy_distribution")
                        if isinstance(pol_dist, pd.DataFrame) and not pol_dist.empty:
                            cl_sheets.append(("Policy Distribution", pol_dist))

                        cop_dist = comp_summary.get("copyleft_distribution")
                        if isinstance(cop_dist, pd.DataFrame) and not cop_dist.empty:
                            cl_sheets.append(("Copyleft Distribution", cop_dist))

                        self.xlsx_renderer.render_multi_sheet(cl_sheets, xlsx_path)
                    else:
                        self.xlsx_renderer.render(table_data, xlsx_path, recipe.name)
                elif has_scan_quality_detail:
                    sheets = [("Summary", table_data)]
                    if not scan_quality_detail_df.empty:
                        sheets.append(("Version Detail", scan_quality_detail_df))
                    self.xlsx_renderer.render_multi_sheet(sheets, xlsx_path)
                else:
                    self.xlsx_renderer.render(table_data, xlsx_path, recipe.name)
                self.logger.debug(f"Generated XLSX: {xlsx_path}")
                generated_files.append(str(xlsx_path))

            # Generate additional raw data files if available (for scan analysis)
            raw_data = report_data.metadata.get("additional_data", {}).get("raw_data")
            if raw_data is not None and hasattr(raw_data, "shape"):
                self.logger.debug(
                    f"Generating additional raw data files with {len(raw_data)} records"
                )

                # CSV raw data output
                if "csv" in formats:
                    raw_csv_path = output_dir / f"{base_filename}_Raw_Data.csv"
                    self.csv_renderer.render(raw_data, raw_csv_path)
                    self.logger.debug(f"Generated Raw Data CSV: {raw_csv_path}")
                    generated_files.append(str(raw_csv_path))
                # XLSX raw data output
                if "xlsx" in formats:
                    raw_xlsx_path = output_dir / f"{base_filename}_Raw_Data.xlsx"
                    self.xlsx_renderer.render(
                        raw_data, raw_xlsx_path, f"{recipe.name} - Raw Data"
                    )
                    self.logger.debug(f"Generated Raw Data XLSX: {raw_xlsx_path}")
                    generated_files.append(str(raw_xlsx_path))
        except Exception as e:
            self.logger.error(f"Error generating table formats: {e}")
            raise
        return generated_files

    def _write_vc_progression_xlsx(self, df: pd.DataFrame, output_path: Path) -> None:
        """Write VC progression DataFrame to XLSX using openpyxl.

        Failed rows (status == 'fetch_failed') get light-red cell fill and
        their numeric columns are replaced with the literal string
        'fetch failed' so the workbook is honest about which versions
        could not be fetched.

        Columns whose values contain lists or dicts (findings_in_version,
        fixed_findings, new_findings, component_churn, externally_changed,
        external_changes_window, etc.) are dropped from the XLSX because
        openpyxl rejects non-scalar cell values. The full nested data is
        still available in the sibling _Progression.csv file.
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        FAILED_FILL = PatternFill(
            start_color="FFF2F2", end_color="FFF2F2", fill_type="solid"
        )
        NUMERIC_COLS = frozenset(
            {"total", "critical", "high", "medium", "low", "components", "new", "fixed"}
        )

        # Drop columns containing non-scalar values (lists/dicts). openpyxl
        # only accepts scalars in cells; the CSV output keeps these columns
        # via pandas' repr-based stringification.
        scalar_df = df.copy()
        for col in list(scalar_df.columns):
            sample_non_null = scalar_df[col].dropna()
            if sample_non_null.empty:
                continue
            first_val = sample_non_null.iloc[0]
            if isinstance(first_val, (list, dict, set, tuple)):
                scalar_df = scalar_df.drop(columns=[col])

        wb = Workbook()
        ws = wb.active
        ws.title = "Progression"

        columns = list(scalar_df.columns)
        ws.append(columns)

        for _, series in scalar_df.iterrows():
            is_failed = series.get("status") == "fetch_failed"
            row_values: list[Any] = []
            for col in columns:
                val = series.get(col)
                if is_failed and col in NUMERIC_COLS:
                    row_values.append("fetch failed")
                else:
                    row_values.append(val)
            ws.append(row_values)
            if is_failed:
                for cell in ws[ws.max_row]:
                    cell.fill = FAILED_FILL

        wb.save(output_path)

    def _render_chart_formats(
        self, recipe: Recipe, report_data: ReportData, output_dir: Path
    ) -> list[str]:
        """Render chart-based formats (HTML only). Returns list of generated file paths."""
        generated_files = []
        try:
            # HTML output
            html_path = output_dir / f"{self._sanitize_filename(recipe.name)}.html"
            self.html_renderer.render(recipe, report_data, html_path)
            self.logger.debug(f"Generated HTML: {html_path}")
            generated_files.append(str(html_path))
        except Exception as e:
            self.logger.error(f"Error generating chart formats: {e}")
            raise
        return generated_files

    def _render_json(
        self,
        recipe: Recipe,
        report_data: ReportData,
        output_dir: Path,
    ) -> list[str]:
        """Render JSON output. Returns list of generated file paths.

        Handles two cases:
        1. Recipes with a dedicated ``json_package`` (e.g. Remediation Package)
        2. Generic table-based recipes — wraps DataFrame as JSON with metadata
        """
        generated_files = []
        try:
            additional_data = report_data.metadata.get("additional_data", {})
            json_package = additional_data.get("json_package")
            if json_package is None:
                tr = additional_data.get("transform_result", {})
                json_package = tr.get("json_package") if isinstance(tr, dict) else None

            base_filename = self._sanitize_filename(recipe.name)

            if json_package and isinstance(json_package, dict):
                # Case 1: dedicated json_package (Remediation Package, etc.)
                json_path = output_dir / f"{base_filename}.json"
                json_path.write_text(
                    json.dumps(json_package, indent=2, default=str),
                    encoding="utf-8",
                )
                self.logger.debug(f"Generated JSON: {json_path}")
                generated_files.append(str(json_path))
            elif (
                isinstance(report_data.data, pd.DataFrame)
                and not report_data.data.empty
            ):
                # Case 2: generic DataFrame → JSON with metadata wrapper
                from datetime import datetime

                metadata: dict[str, Any] = {
                    "recipe": recipe.name,
                    "generated_at": datetime.now(UTC).isoformat(),
                }
                if self.config:
                    if getattr(self.config, "project_filter", None):
                        metadata["project"] = self.config.project_filter
                    if getattr(self.config, "folder_filter", None):
                        metadata["folder"] = self.config.folder_filter
                    filters: dict[str, str] = {}
                    if getattr(self.config, "component_filter", None):
                        filters["component"] = self.config.component_filter
                    if getattr(self.config, "cve_filter", None):
                        filters["cve"] = self.config.cve_filter
                    if filters:
                        metadata["filters"] = filters

                records = report_data.data.to_dict(orient="records")
                # Replace float NaN with None for valid JSON output
                for rec in records:
                    for k, v in rec.items():
                        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                            rec[k] = None
                output_obj = {
                    "metadata": metadata,
                    "data": records,
                    "count": len(records),
                }
                json_path = output_dir / f"{base_filename}.json"
                json_path.write_text(
                    json.dumps(output_obj, indent=2, default=str),
                    encoding="utf-8",
                )
                self.logger.debug(f"Generated JSON: {json_path}")
                generated_files.append(str(json_path))
        except Exception as e:
            self.logger.error(f"Error generating JSON: {e}")
            raise
        return generated_files

    def _render_pdf(
        self,
        recipe: Recipe,
        report_data: ReportData,
        output_dir: Path,
    ) -> list[str]:
        """Render PDF output via weasyprint. Returns list of generated file paths."""
        generated_files = []
        try:
            from fs_report.renderers.pdf_renderer import PDFRenderer

            pdf_renderer = PDFRenderer()
            pdf_path = output_dir / f"{self._sanitize_filename(recipe.name)}.pdf"
            pdf_renderer.render(recipe, report_data, pdf_path)
            self.logger.debug(f"Generated PDF: {pdf_path}")
            generated_files.append(str(pdf_path))
        except Exception as e:
            self.logger.error(f"Error generating PDF: {e}")
            raise
        return generated_files

    def _render_markdown(
        self,
        recipe: Recipe,
        report_data: ReportData,
        output_dir: Path,
    ) -> list[str]:
        """Render agent-optimized Markdown report. Returns list of generated file paths."""
        generated_files = []
        try:
            base_filename = self._sanitize_filename(recipe.name)
            md_path = output_dir / f"{base_filename}.md"
            self.md_renderer.render(recipe, report_data, md_path)
            generated_files.append(str(md_path))
        except Exception as e:
            self.logger.error(f"Error generating Markdown: {e}")
            raise
        return generated_files

    def _sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for safe file system usage."""
        # Replace problematic characters
        sanitized = filename.replace("/", "_").replace("\\", "_")
        sanitized = sanitized.replace(":", "_").replace("*", "_")
        sanitized = sanitized.replace("?", "_").replace('"', "_")
        sanitized = sanitized.replace("<", "_").replace(">", "_")
        sanitized = sanitized.replace("|", "_")

        # Remove leading/trailing spaces and dots
        sanitized = sanitized.strip(" .")

        # Ensure filename is not empty
        if not sanitized:
            sanitized = "report"

        return sanitized
